"""
admin/reports.py
Reports: summary stats + export to PDF, Excel, CSV
"""

import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))


import customtkinter as ctk
import tkinter as tk
import os
import datetime
from config.database import Database
from theme import COLORS, FONTS, make_button, make_label, make_card, show_message


REPORTS_DIR = os.path.join(os.path.dirname(__file__), "..", "reports")


class ReportsPanel(ctk.CTkFrame):
    def __init__(self, parent, **kwargs):
        super().__init__(parent, fg_color=COLORS["bg_dark"], corner_radius=0, **kwargs)
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self._build()

    def _build(self):
        make_label(self, "Reports & Analytics", style="heading").grid(
            row=0, column=0, sticky="w", padx=24, pady=(20, 16))

        scroll = ctk.CTkScrollableFrame(self, fg_color=COLORS["bg_dark"],
                                        scrollbar_button_color=COLORS["bg_card"])
        scroll.grid(row=1, column=0, sticky="nsew", padx=24, pady=(0, 16))
        scroll.grid_columnconfigure((0, 1), weight=1)

        # Movie report card
        self._movie_card(scroll, row=0)
        # User report card
        self._user_card(scroll, row=0, col=1)
        # Export section
        self._export_section(scroll, row=1)

    def _movie_card(self, parent, row=0, col=0):
        card = make_card(parent)
        card.grid(row=row, column=col, padx=8, pady=8, sticky="nsew")
        make_label(card, "🎬  Movie Report", style="subhead").pack(anchor="w", padx=20, pady=(16, 12))

        stats = self._get_movie_stats()
        for label, val in stats:
            row_ = ctk.CTkFrame(card, fg_color="transparent")
            row_.pack(fill="x", padx=20, pady=3)
            make_label(row_, label, style="small").pack(side="left")
            ctk.CTkLabel(row_, text=str(val), font=FONTS["body"],
                         text_color=COLORS["accent"]).pack(side="right")

        make_label(card, "Top Rated Movies", style="small").pack(anchor="w", padx=20, pady=(12, 4))
        top = list(Database.movies().find().sort("rating", -1).limit(5))
        for m in top:
            row_ = ctk.CTkFrame(card, fg_color=COLORS["bg_input"], corner_radius=6)
            row_.pack(fill="x", padx=20, pady=2)
            ctk.CTkLabel(row_, text=m.get("title","")[:30], font=FONTS["small"],
                         text_color=COLORS["text_primary"]).pack(side="left", padx=8, pady=4)
            ctk.CTkLabel(row_, text=f"⭐ {m.get('rating','')}", font=FONTS["small"],
                         text_color=COLORS["accent"]).pack(side="right", padx=8)

        ctk.CTkFrame(card, fg_color="transparent", height=16).pack()

    def _user_card(self, parent, row=0, col=1):
        card = make_card(parent)
        card.grid(row=row, column=col, padx=8, pady=8, sticky="nsew")
        make_label(card, "👥  User Report", style="subhead").pack(anchor="w", padx=20, pady=(16, 12))

        for label, val in self._get_user_stats():
            row_ = ctk.CTkFrame(card, fg_color="transparent")
            row_.pack(fill="x", padx=20, pady=3)
            make_label(row_, label, style="small").pack(side="left")
            ctk.CTkLabel(row_, text=str(val), font=FONTS["body"],
                         text_color=COLORS["info"]).pack(side="right")

        ctk.CTkFrame(card, fg_color="transparent", height=16).pack()

    def _export_section(self, parent, row=1):
        card = make_card(parent)
        card.grid(row=row, column=0, columnspan=2, padx=8, pady=8, sticky="ew")
        make_label(card, "📤  Export Reports", style="subhead").pack(anchor="w", padx=20, pady=(16, 12))

        btn_row = ctk.CTkFrame(card, fg_color="transparent")
        btn_row.pack(padx=20, pady=(0, 20))

        make_button(btn_row, "📄  Export PDF",   command=self._export_pdf,   width=160).pack(side="left", padx=6)
        make_button(btn_row, "📊  Export Excel", command=self._export_excel, width=160).pack(side="left", padx=6)
        make_button(btn_row, "📋  Export CSV",   command=self._export_csv,   width=160).pack(side="left", padx=6)

        self._status_lbl = make_label(card, "", style="small")
        self._status_lbl.pack(anchor="w", padx=20, pady=(0, 12))

    # ─── Stats helpers ─────────────────────────────────────────────────────────
    def _get_movie_stats(self):
        total   = Database.movies().count_documents({})
        genres  = Database.genres().count_documents({})
        ratings = [float(m.get("rating",0) or 0) for m in Database.movies().find()]
        avg     = sum(ratings)/len(ratings) if ratings else 0
        return [("Total Movies", total), ("Total Genres", genres),
                ("Average Rating", f"{avg:.2f}"), ("Total Reviews", Database.reviews().count_documents({}))]

    def _get_user_stats(self):
        total   = Database.users().count_documents({})
        active  = Database.users().count_documents({"status": "active"})
        blocked = Database.users().count_documents({"status": "blocked"})
        return [("Total Users", total), ("Active Users", active), ("Blocked Users", blocked)]

    # ─── Export Methods ────────────────────────────────────────────────────────
    def _ts(self):
        return datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

    def _export_pdf(self):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors as rl_colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet

            os.makedirs(REPORTS_DIR, exist_ok=True)
            path = os.path.join(REPORTS_DIR, f"report_{self._ts()}.pdf")

            doc    = SimpleDocTemplate(path, pagesize=A4)
            styles = getSampleStyleSheet()
            elems  = []

            elems.append(Paragraph("CineVault – Management Report", styles["Title"]))
            elems.append(Spacer(1, 12))
            elems.append(Paragraph(f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
            elems.append(Spacer(1, 20))

            # Movie table
            elems.append(Paragraph("Movies", styles["Heading2"]))
            movies = list(Database.movies().find())
            data   = [["Title", "Genre", "Director", "Year", "Rating"]]
            for m in movies:
                data.append([m.get("title",""), m.get("genre",""), m.get("director",""),
                              str(m.get("year","")), str(m.get("rating",""))])
            t = Table(data, colWidths=[130, 80, 100, 50, 60])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), rl_colors.HexColor("#E8A020")),
                ("TEXTCOLOR",  (0,0), (-1,0), rl_colors.black),
                ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0,1), (-1,-1), [rl_colors.HexColor("#F5F5F5"), rl_colors.white]),
                ("GRID", (0,0), (-1,-1), 0.5, rl_colors.grey),
            ]))
            elems.append(t)
            elems.append(Spacer(1, 20))

            # Users table
            elems.append(Paragraph("Users", styles["Heading2"]))
            users = list(Database.users().find())
            ud = [["Name", "Username", "Email", "Role", "Status"]]
            for u in users:
                ud.append([u.get("name",""), u.get("username",""), u.get("email",""),
                           u.get("role",""), u.get("status","")])
            ut = Table(ud, colWidths=[110, 80, 140, 60, 60])
            ut.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), rl_colors.HexColor("#3498DB")),
                ("TEXTCOLOR",  (0,0), (-1,0), rl_colors.white),
                ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0,1), (-1,-1), [rl_colors.HexColor("#F5F5F5"), rl_colors.white]),
                ("GRID", (0,0), (-1,-1), 0.5, rl_colors.grey),
            ]))
            elems.append(ut)

            doc.build(elems)
            self._status_lbl.configure(text=f"✓ PDF saved: {path}", text_color=COLORS["success"])
        except Exception as e:
            show_message(self, "Error", f"PDF export failed: {e}", "error")

    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            os.makedirs(REPORTS_DIR, exist_ok=True)
            path = os.path.join(REPORTS_DIR, f"report_{self._ts()}.xlsx")

            wb = openpyxl.Workbook()

            # Movies sheet
            ws = wb.active
            ws.title = "Movies"
            headers = ["Title", "Genre", "Director", "Year", "Language", "Rating", "Description"]
            self._xl_header(ws, headers)
            for m in Database.movies().find():
                ws.append([m.get("title",""), m.get("genre",""), m.get("director",""),
                           str(m.get("year","")), m.get("language",""),
                           str(m.get("rating","")), m.get("description","")])

            # Users sheet
            ws2 = wb.create_sheet("Users")
            self._xl_header(ws2, ["Name","Username","Email","Role","Status"])
            for u in Database.users().find():
                ws2.append([u.get("name",""), u.get("username",""), u.get("email",""),
                            u.get("role",""), u.get("status","")])

            # Reviews sheet
            ws3 = wb.create_sheet("Reviews")
            self._xl_header(ws3, ["User","Movie","Rating","Review","Date"])
            for r in Database.reviews().find():
                u = Database.users().find_one({"_id": r.get("user_id")})
                m = Database.movies().find_one({"_id": r.get("movie_id")})
                ws3.append([u.get("username","?") if u else "?",
                            m.get("title","?") if m else "?",
                            str(r.get("rating","")), r.get("review",""),
                            str(r.get("date",""))[:10]])

            wb.save(path)
            self._status_lbl.configure(text=f"✓ Excel saved: {path}", text_color=COLORS["success"])
        except Exception as e:
            show_message(self, "Error", f"Excel export failed: {e}", "error")

    def _xl_header(self, ws, headers):
        from openpyxl.styles import Font, PatternFill
        fill = PatternFill("solid", fgColor="E8A020")
        ws.append(headers)
        for cell in ws[1]:
            cell.font   = Font(bold=True)
            cell.fill   = fill

    def _export_csv(self):
        try:
            import csv
            os.makedirs(REPORTS_DIR, exist_ok=True)
            path = os.path.join(REPORTS_DIR, f"movies_{self._ts()}.csv")
            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["Title","Genre","Director","Year","Language","Rating"])
                for m in Database.movies().find():
                    w.writerow([m.get("title",""), m.get("genre",""), m.get("director",""),
                                m.get("year",""), m.get("language",""), m.get("rating","")])
            self._status_lbl.configure(text=f"✓ CSV saved: {path}", text_color=COLORS["success"])
        except Exception as e:
            show_message(self, "Error", f"CSV export failed: {e}", "error")
